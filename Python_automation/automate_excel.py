#we can use python code to automate to process the data in excel sheet

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
#include BarChart and Reference as class so that --> we can use it for generating bar chart
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']


# print(sheet.max_row)


    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)#keep the row 2, 3, 4 and 3rd column that is the price
    #as our work is only with values which start from 2nd row
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        #as we are adding a new col in the sheet , which is rows and 4th col 
        corrected_price_cell.value  = corrected_price

    values  = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col = 4,
            max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)
#let us assume that we want to add the chart after e2 column continously --> so 
#we used the chart object , and used BarChart and used chart.add_data(values) in which 
#we get values as parameters of updates prices 
# e2 is used --> that is E col and th row we want a chart

#create new file 
# 4 --> this is the number of row in the excel sheet 

#we have used xl to use as module . operator in it to just simplify the code
#xl is called as alias 
#to work with excel sheet we have a function called loadworkbook use it and pass the ar as excel sheet name

#openpyxl is module or package used to work with the excel spreadsheets
#we know that there is only one sheet inside the excel file , so specify the sheet1 as indexing in the list 

#we stored the value in wb object
#return the first sheet in the excel which is sheet1 be -> careful as Sheet1 name is case sensitive


#In excel sheet we have various cells called as data as combination of row and column 
#by using coordinates of the cell --> that is combination of rows and columns 

#there are 2 methods to access the cell

#1. by specifying the name of cell that is --> A column and 1st row so a1 

#2. by using cell method with . operator in sheet --> cell = sheet.cell(1,1) specify the row and col


# we can use sheet.max_row function to get maximum rows present in the excel sheet
